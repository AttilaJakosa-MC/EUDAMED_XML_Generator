import os
import pandas as pd
import warnings
import re
import oracledb
from sqlalchemy import create_engine
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- Configuration ---
# Update your database credentials and alias here before running
db_user = os.getenv("ORAUSER")
db_password = os.getenv("ORAPW")
db_alias = "MC9ELES"  # The network alias from your tnsnames.ora file

BASE_DIR = r"K:\PMO Projects\24_01_E New ERP Introduction\Migration\Eudamed"
SQL_FILE_PATH = os.path.join(BASE_DIR, r"plsql\Create EUDAMED table.plsql")

# Save outputs to the user's Downloads folder
DOWNLOADS_DIR = os.path.join(os.path.expanduser('~'), 'Downloads')
OUTPUT_EUDAMED = os.path.join(DOWNLOADS_DIR, "EUDAMED.xlsx")
OUTPUT_CONFIG = os.path.join(DOWNLOADS_DIR, "EUDAMED_config.xlsx")

def check_for_special_characters(df):
    """
    Checks for strictly problematic invisible formatting characters 
    (e.g., non-breaking spaces \xA0, zero-width spaces, smart quotes, control chars).
    Ignores valid Central European (é, á, ű, etc.), Russian (Cyrillic), and standard text.
    """
    # Specifically target known VBA-breaking invisible or formatting characters
    # \u00A0 : Non-breaking space (causing 1004)
    # \u200B-\u200F : Zero-width spaces and markers
    # \u201C-\u201D : Left/Right double quotation marks (smart quotes)
    # \u2018-\u2019 : Left/Right single quotation marks
    # \u0000-\u0008, \u000B, \u000C, \u000D, \u000E, \u000F-\u001F : Unprintable control characters (excluding tab \t, newline \n)
    # Note: we are including \r (\u000D) to be cleaned.
    bad_chars_regex = r'[\u00A0\u200B-\u200F\u2028\u2029\u201C\u201D\u2018\u2019\u0000-\u0008\u000B-\u000F\u0010-\u001F]'
    special_char_pattern = re.compile(bad_chars_regex)
    
    problematic_columns = []
    
    # Track the detailed report
    warning_details = []

    for col in df.select_dtypes(include=['object', 'string']).columns:
        # Find exactly which rows have the issue
        mask = df[col].fillna("").astype(str).str.contains(special_char_pattern, regex=True)
        if mask.any():
            problematic_columns.append(col)
            
            # Extract bad rows
            bad_rows = df[mask]
            
            warning_details.append(f"\n--- Column: '{col}' ---")
            
            # Show a sample of the bad rows (first 5 per column so we don't flood the console)
            for idx, row_series in bad_rows.head(5).iterrows():
                val = str(row_series[col])
                
                # Find all exact bad characters in this specific string
                bad_chars_found = list(set(re.findall(special_char_pattern, val)))
                char_codes = [f"Chr({ord(c)})" for c in bad_chars_found]
                
                # We can also attempt to pull the PCODE or PARTNO if it exists to identify the row easily
                row_identifier = f"Row Index {idx}"
                if 'PCODE' in df.columns:
                    row_identifier += f" | PCODE: {row_series.get('PCODE', 'N/A')}"
                elif 'PARTNO' in df.columns:
                    row_identifier += f" | PARTNO: {row_series.get('PARTNO', 'N/A')}"
                
                warning_details.append(
                    f"  {row_identifier}\n"
                    f"  Value: '{val}'\n"
                    f"  Bad Characters Hidden In Value: {', '.join(char_codes)}"
                )
            
            if len(bad_rows) > 5:
                warning_details.append(f"  ... and {len(bad_rows) - 5} more rows in this column.")
            
    if problematic_columns:
        report = "\n".join(warning_details)
        warnings.warn(
            f"\n\n*** NOTICE ***\nInvisible or special UTF-8 characters were detected in the source data for columns: {problematic_columns}\n"
            f"Details of first few offending rows:\n{report}\n\n"
            f"Don't worry, the script will automatically clean these up before exporting to Excel."
        )

def clean_special_characters(df):
    """
    Cleans problematic invisible or formatting characters from the DataFrame.
    Replaces non-breaking spaces with normal spaces, smart quotes with straight quotes,
    and removes zero-width spaces and unprintable control characters.
    """
    def clean_text(text):
        if not isinstance(text, str):
            return text
        text = re.sub(r'\u00A0', ' ', text)
        text = re.sub(r'[\u201C\u201D]', '"', text)
        text = re.sub(r'[\u2018\u2019]', "'", text)
        text = re.sub(r'[\u200B-\u200F\u0000-\u0008\u000B-\u000F\u0010-\u001F\u2028\u2029]', '', text)
        return text

    for col in df.select_dtypes(include=['object', 'string']).columns:
        df[col] = df[col].map(clean_text)

def write_df_to_excel_table(df, filepath, table_name):
    """
    Writes a Pandas DataFrame to an Excel file, formatted as an official Excel Table (ListObject)
    where all cells have 'Shrink to Fit' enabled.
    """
    # Excel tables require unique, non-empty header strings (no spaces perfectly mapped)
    # This prevents openpyxl from failing if a column name is totally blank
    clean_columns = []
    seen = set()
    for i, c in enumerate(df.columns):
        c_str = str(c).strip()
        if not c_str:
            c_str = f"Col_{i}"
        # Ensure uniqueness
        if c_str in seen:
            c_str = f"{c_str}_{i}"
        seen.add(c_str)
        clean_columns.append(c_str)
    
    df.columns = clean_columns

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        worksheet = writer.sheets['Data']
        
        max_row = len(df) + 1  # Include the header row
        max_col = len(df.columns)
        
        if max_row > 1 and max_col > 0:
            # Create a table range like A1:Z100
            end_col_letter = get_column_letter(max_col)
            table_range = f"A1:{end_col_letter}{max_row}"
            
            # Formally declare it as an Excel Table
            tab = Table(displayName=table_name, ref=table_range)
            # Add a basic blue table style so it looks nice automatically
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
        
        # Apply Shrink-to-Fit alignment to absolutely every cell
        shrink_alignment = Alignment(shrink_to_fit=True)
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = shrink_alignment
                
                # If a string starts with `'=`, keep the `'` but force text behavior
                if isinstance(cell.value, str) and cell.value.startswith("'="):
                    cell.data_type = 's'
                    # cell.value is untouched, so it keeps the apostrophe

def main():
    print(f"Reading SQL file from: {SQL_FILE_PATH}")
    if not os.path.exists(SQL_FILE_PATH):
        print("Error: SQL file does not exist at the specified path.")
        return
        
    with open(SQL_FILE_PATH, 'r', encoding='utf-8') as f:
        sql_query = f.read()

    # Pre-process the SQL script:
    # 1. Remove SQL*Plus specific commands (like SET DEFINE OFF;)
    sql_query = re.sub(r'^\s*SET\s+.*?;', '', sql_query, flags=re.IGNORECASE | re.MULTILINE)
    
    # 2. Extract ALTER SESSION statements to run them separately before the main query
    alter_session_matches = re.finditer(r'^\s*(ALTER\s+SESSION\s+.*?);', sql_query, flags=re.IGNORECASE | re.MULTILINE)
    alter_statements = [m.group(1) for m in alter_session_matches]
    
    # 3. Remove ALTER SESSION statements from the main query body
    sql_query = re.sub(r'^\s*ALTER\s+SESSION\s+.*?;', '', sql_query, flags=re.IGNORECASE | re.MULTILINE)
    
    # 4. Remove all comments from the entire query string
    # Python tools often freak out if the query ends in a comment block instead of the final ORDER BY.
    # Block comments (/* ... */)
    sql_query = re.sub(r'/\*.*?\*/', '', sql_query, flags=re.DOTALL)
    
    # Inline comments (-- ...)
    sql_query = re.sub(r'--.*?$', '', sql_query, flags=re.MULTILINE)
    
    # 5. Clean up any trailing whitespace, and remove the single mandatory trailing semicolon 
    # (SQLAlchemy hates trailing semicolons for SELECTs)
    sql_query = sql_query.strip()
    if sql_query.endswith(';'):
        sql_query = sql_query[:-1]
    sql_query = sql_query.strip()

    print(f"Connecting to the database via alias '{db_alias}'...")
    try:
        from sqlalchemy import text
        
        # Establish DB connection using SQLAlchemy (matching the DPT scripts method)
        engine = create_engine(f"oracle+oracledb://{db_user}:{db_password}@{db_alias}")
        
        with engine.connect() as conn:
            # Execute session setup parameters first
            for alt_stmt in alter_statements:
                print(f"Executing Session Parameter: {alt_stmt}")
                conn.execute(text(alt_stmt))
            
            print("Executing query and fetching data... (This may take a moment depending on query complexity)")
            # pandas read_sql will use the connection to automatically manage fetching
            df = pd.read_sql(text(sql_query), conn)
        
    except Exception as e:
        import traceback
        
        # SQLAlchemy wraps errors and appends the massive SQL text after "[SQL:" 
        # We can split that off so you only see the actual Oracle error without spamming the console.
        err_msg = str(e)
        if "[SQL:" in err_msg:
            err_msg = err_msg.split("[SQL:")[0].strip()
            
        print(f"\n--- ERROR ENCOUNTERED ---")
        print(f"Type: {type(e).__name__}")
        print(f"Message: {err_msg}")
        
        print("\nWhere it happened (Python Stack):")
        tb = traceback.extract_tb(e.__traceback__)
        for frame in tb:
            # Only print our own script's stack trace context to keep it clean
            if "MGEN table creator" in frame.filename:
                print(f"  -> Line {frame.lineno} in {frame.name}")
        
        print("\nPlease check the error message above.")
        return
    finally:
        if 'engine' in locals():
            engine.dispose()
            
    print(f"Data fetched successfully. Total rows: {len(df)}")
    
    # Check for invisible/special characters to prevent future Excel issues
    check_for_special_characters(df)
    
    # Clean up those special characters automatically 
    print("Cleaning special characters...")
    clean_special_characters(df)
    
    # Handle possible case-sensitivity issues from the DB column names
    upper_cols = {c.upper(): c for c in df.columns}
    if 'NAME' in upper_cols:
        # Rename physical column to exactly "NAME" to avoid KeyErrors
        df.rename(columns={upper_cols['NAME']: 'NAME'}, inplace=True)
    else:
        print("Error: 'NAME' column not found in the result set!")
        print(f"Available columns: {list(df.columns)}")
        return

    # Filter data into two DataFrames
    print("Filtering data...")
    config_names = [
        'DEVICE_CREATE_ENV', 'BASIC_UDI_UPDATE_ENV','UDI_DI_ENV', 'UDI_DI_LIMIT',
        'DEVICE_BASIC_UDI_CREATE_PAYLOAD_ROOT',
        'DEVICE_UDI_DI_CREATE_PAYLOAD_ROOT', 'BASIC_UDI_UPDATE_PAYLOAD_ROOT',
        'UDI_DI_PAYLOAD_ROOT', 'XML_OBJECT_ORDER', 'XSD_VERSION'
    ]
    is_config = df['NAME'].isin(config_names)
    
    df_config = df[is_config].copy()
    df_eudamed = df[~is_config].copy()
    
    # Export to Excel as styled ListObject Tables with Shrink to Fit
    print(f"Exporting exactly {len(df_config)} rows to {OUTPUT_CONFIG}")
    write_df_to_excel_table(df_config, OUTPUT_CONFIG, "CONFIG_TABLE")
    
    print(f"Exporting exactly {len(df_eudamed)} rows to {OUTPUT_EUDAMED}")
    write_df_to_excel_table(df_eudamed, OUTPUT_EUDAMED, "EUDAMED_TABLE")
    
    print("Data export complete!")

if __name__ == "__main__":
    main()
